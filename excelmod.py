import openpyxl

def row_count(path,s):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(s)
    return sheet.max_row
def col_count(path,s):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(s)
    return sheet.max_column
def read(path,s,r,c):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(s)
    return sheet.cell(row=r,column=c).value
def write(path,s,r,c,data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(s)
    sheet.cell(row=r,column=c).value=data
    workbook.save(path)